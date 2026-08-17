"""Workbook loading and row projection.

The reconstruction reasons jointly over the whole row: the fix explains what
the question was really asking, and the question explains what the fix is for.
``CaseView`` is therefore the full permitted context for one case, assembled
once and passed to the model unchanged.
"""
from __future__ import annotations

import re
import warnings
from dataclasses import dataclass, field
from typing import Iterator

import openpyxl

from . import config

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

REQUIRED_COLUMNS = (
    "platform",
    "issue_number",
    "issue_id",
    "title",
    "category",
    "buggy_code",
    "fixed_code",
    "buggy_question_description",
    "fixed_solution_explanation",
)

#: Excel caps a single cell at this many characters; longer source text was
#: silently truncated before it ever reached this workbook.
EXCEL_CELL_LIMIT = 32767


class SourceIntegrityError(RuntimeError):
    """Raised when the workbook does not match documented expectations."""


@dataclass(frozen=True)
class CaseView:
    """The full context handed to the reconstruction step for one case."""

    case_id: str
    issue_number: int
    issue_id: str
    platform: str
    title: str
    category_label: str
    buggy_code: str
    buggy_question_description: str
    fixed_code: str
    fixed_solution_explanation: str


@dataclass
class SourceRow:
    """One spreadsheet row, verbatim."""

    platform: str
    issue_number: int
    issue_id: str
    title: str
    category: str
    buggy_code: str
    fixed_code: str
    buggy_question_description: str
    fixed_solution_explanation: str
    excel_row: int
    dir_name: str = ""
    truncated_fields: tuple = field(default_factory=tuple)

    def view(self) -> CaseView:
        return CaseView(
            case_id=self.dir_name,
            issue_number=self.issue_number,
            issue_id=self.issue_id,
            platform=self.platform,
            title=self.title.strip(),
            category_label=sanitize_category_label(self.category),
            buggy_code=self.buggy_code,
            buggy_question_description=self.buggy_question_description,
            fixed_code=self.fixed_code,
            fixed_solution_explanation=self.fixed_solution_explanation,
        )


# --------------------------------------------------------------------------
# category normalisation
# --------------------------------------------------------------------------
def sanitize_category_label(raw: str) -> str:
    """Reduce the noisy ``category`` cell to a bare taxonomy label.

    79% of non-empty ``category`` values append a verdict essay ("Why AI Wins
    in This Category: ..."). That tail is about which answer source was better,
    not about what kind of bug this is, so it is trimmed to keep the prompt on
    topic. Returns '' when nothing usable survives.
    """
    if not raw:
        return ""
    text = str(raw).strip()
    if not text or text.lower() == "none":
        return ""

    text = re.sub(r"^\s*category\s*:\s*", "", text, flags=re.I)

    lowered = text.lower()
    cut = len(text)
    for marker in config.CATEGORY_LEAK_MARKERS:
        pos = lowered.find(marker)
        if pos != -1:
            cut = min(cut, pos)
    text = text[:cut]

    for line in text.splitlines():
        line = line.strip().strip(":").strip()
        if line and line.lower() != "none":
            return line[: config.CATEGORY_LABEL_MAX_CHARS].strip()
    return ""


# --------------------------------------------------------------------------
# Directory naming
# --------------------------------------------------------------------------
def _pad(n: int) -> str:
    return f"{n:0{config.PAD_WIDTH}d}"


def assign_dir_names(rows: list[SourceRow]) -> None:
    """Populate ``dir_name`` on every row per ``config.NAMING_SCHEME``.

    ``issue_number`` alone is not unique (531 distinct values across 556 rows),
    so the default scheme disambiguates only the colliding numbers and leaves
    the other 506 as the plain ``issue_NNN`` form.
    """
    scheme = config.NAMING_SCHEME
    counts: dict[int, int] = {}
    for r in rows:
        counts[r.issue_number] = counts.get(r.issue_number, 0) + 1

    for r in rows:
        n = _pad(r.issue_number)
        plat = r.platform.strip().lower() or "na"
        if scheme == "number_only":
            r.dir_name = f"issue_{n}"
        elif scheme == "prefix_platform":
            r.dir_name = f"issue_{plat}_{n}"
        elif scheme == "append_issue_id":
            r.dir_name = f"issue_{n}__{r.issue_id}"
        elif scheme == "suffix_collisions":
            r.dir_name = f"issue_{n}" if counts[r.issue_number] == 1 else f"issue_{n}_{plat}"
        else:
            raise ValueError(f"unknown NAMING_SCHEME {scheme!r}")

    seen: dict[str, str] = {}
    for r in rows:
        if r.dir_name in seen and scheme != "number_only":
            raise SourceIntegrityError(
                f"naming scheme {scheme!r} produced a collision: {r.dir_name} "
                f"(issue_id {seen[r.dir_name]} and {r.issue_id})"
            )
        seen[r.dir_name] = r.issue_id


# --------------------------------------------------------------------------
# Loading
# --------------------------------------------------------------------------
def _cell(value) -> str:
    return "" if value is None else str(value)


def load_rows(workbook_path=None, sheet_name: str | None = None) -> list[SourceRow]:
    """Read the confirmed sheet. Refuses to read any other sheet."""
    path = workbook_path or config.WORKBOOK
    sheet = sheet_name or config.SOURCE_SHEET

    if sheet in config.FORBIDDEN_SHEETS:
        raise SourceIntegrityError(
            f"sheet {sheet!r} is explicitly excluded as a data source; "
            f"only {config.SOURCE_SHEET!r} may be used"
        )
    if sheet != config.SOURCE_SHEET:
        raise SourceIntegrityError(
            f"refusing to read {sheet!r}; the sole source of truth is {config.SOURCE_SHEET!r}"
        )
    if not path.exists():
        raise SourceIntegrityError(f"workbook not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise SourceIntegrityError(f"worksheet {sheet!r} not present in {path.name}")
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        header = [_cell(h).strip() for h in next(it)]
        idx = {h: i for i, h in enumerate(header)}

        missing = [c for c in REQUIRED_COLUMNS if c not in idx]
        if missing:
            raise SourceIntegrityError(f"missing required columns: {missing}")

        rows: list[SourceRow] = []
        for excel_row, raw in enumerate(it, start=2):
            if not any(c is not None and str(c).strip() for c in raw):
                continue

            def col(name: str) -> str:
                return _cell(raw[idx[name]])

            num_raw = col("issue_number").strip()
            try:
                issue_number = int(float(num_raw))
            except ValueError as exc:
                raise SourceIntegrityError(
                    f"row {excel_row}: issue_number {num_raw!r} is not numeric"
                ) from exc

            trunc = tuple(
                name
                for name in ("buggy_code", "fixed_code", "buggy_question_description",
                             "fixed_solution_explanation")
                if len(col(name)) >= EXCEL_CELL_LIMIT
            )

            rows.append(
                SourceRow(
                    platform=col("platform").strip(),
                    issue_number=issue_number,
                    issue_id=col("issue_id").strip(),
                    title=col("title"),
                    category=col("category"),
                    buggy_code=col("buggy_code"),
                    fixed_code=col("fixed_code"),
                    buggy_question_description=col("buggy_question_description"),
                    fixed_solution_explanation=col("fixed_solution_explanation"),
                    excel_row=excel_row,
                    truncated_fields=trunc,
                )
            )
    finally:
        wb.close()

    if len(rows) != config.EXPECTED_ROW_COUNT:
        raise SourceIntegrityError(
            f"expected {config.EXPECTED_ROW_COUNT} data rows, found {len(rows)}"
        )

    assign_dir_names(rows)
    return rows


def iter_rows() -> Iterator[SourceRow]:
    yield from load_rows()
